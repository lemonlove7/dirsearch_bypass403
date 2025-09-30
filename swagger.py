# -*- coding: utf-8 -*-

import json
import sys
import os
import argparse
import requests
import re
import random
import time
import urllib3
from datetime import datetime
from urllib.parse import urlparse, urljoin, parse_qs
from loguru import logger

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
except ImportError:
    logger.error("Selenium library not found. Please run 'pip install selenium' to install it.")
    sys.exit(1)
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    logger.error("openpyxl library not found. Please run 'pip install openpyxl' to install it.")
    sys.exit(1)


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
logger.remove(); logger.add(sys.stderr, level="INFO")
logger.add("debug.log", level="DEBUG", format="{time:YYYY-MM-DD HH:mm:ss} | {level: <8} | {message}")
proxies = {'https': 'http://127.0.0.1:7890', 'http': 'http://127.0.0.1:7890'}
SET_PROXY = False
header_agents = [
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36'
]
ALLOWED_METHODS = ['get', 'post']
DANGEROUS_KEYWORDS = [
    'delete', 'remove', 'destroy', 'drop', 'del', 'erase', 'update', 'modify', 'edit', 'set', 'change', 'upd',
    'mod', 'patch', 'put', 'add', 'create', 'new', 'insert', 'save', 'upload'
]
PATH_VARIABLE_FORMAT = "{{{param}}}"


workbook = Workbook()
if "Sheet" in workbook.sheetnames: workbook.remove(workbook["Sheet"])
ws_all_apis = workbook.create_sheet(title="所有API")
ws_called_apis = workbook.create_sheet(title="已调用API")
ws_filtered_apis = workbook.create_sheet(title="已过滤API")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="DengXian"); DATA_FONT = Font(name="DengXian")
HEADER_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILTER_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def setup_xlsx_headers():
    headers_all = ["源API文档", "请求方法", "请求URL", "接口摘要", "状态"]
    headers_called = ["源API文档", "请求方法", "请求URL", "接口摘要", "请求参数", "请求头", "状态码", "响应内容"]  # 新增“请求头”列
    headers_filtered = ["源API文档", "请求方法", "请求URL", "接口摘要", "过滤原因"]
    sheets_headers = {ws_all_apis: headers_all, ws_called_apis: headers_called, ws_filtered_apis: headers_filtered}
    for ws, headers in sheets_headers.items():
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = HEADER_ALIGNMENT; cell.border = THIN_BORDER

def apply_row_styles(worksheet, row_index, fill=None):
    for cell in worksheet[row_index]:
        cell.font = DATA_FONT; cell.alignment = CELL_ALIGNMENT; cell.border = THIN_BORDER
        if fill: cell.fill = fill

def auto_adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0; column = col[0].column_letter
        for cell in col:
            try:
                cell_len = 0
                if cell.value:
                    for char in str(cell.value): cell_len += 2 if '\u4e00' <= char <= '\u9fff' else 1
                if cell_len > max_length: max_length = cell_len
            except: pass
        adjusted_width = max_length + 2
        if adjusted_width > 80: adjusted_width = 80
        if adjusted_width < 12: adjusted_width = 12
        worksheet.column_dimensions[column].width = adjusted_width

def is_url_dangerous(url, keywords):
    try:
        path = urlparse(url).path.lower()
        for keyword in keywords:
            if keyword in path:
                reason = f"URL路径包含危险关键词: '{keyword}'"; logger.warning(f"{reason}. 将跳过请求."); return True, reason
        return False, None
    except Exception:
        reason = f"URL解析失败: {url}"; logger.error(f"{reason}. 为安全起见, 跳过."); return True, reason

# -------------------------- 新增：解析自定义headers（处理key:value格式） --------------------------
def parse_custom_headers(header_list):
    """
    解析命令行输入的header列表（格式：key:value），返回字典格式的headers
    :param header_list: argparse接收的--header参数列表（如 ["token:123456", "User-Agent:Custom"]）
    :return: 解析后的headers字典，无效格式会过滤并警告
    """
    custom_headers = {}
    if not header_list:
        return custom_headers
    
    for header_str in header_list:
        # 分割key和value（只按第一个冒号分割，支持value含冒号的场景）
        if ':' not in header_str:
            logger.warning(f"无效的header格式: '{header_str}'，需符合 'key:value'，已跳过")
            continue
        key, value = header_str.split(':', 1)  # 1表示只分割一次
        key = key.strip()
        value = value.strip()
        # 过滤空键/空值
        if not key:
            logger.warning(f"header键为空: '{header_str}'，已跳过")
            continue
        if not value:
            logger.warning(f"header值为空（键: {key}），已跳过")
            continue
        custom_headers[key] = value
    
    if custom_headers:
        logger.info(f"已加载自定义headers: {json.dumps(custom_headers, ensure_ascii=False)}")
    else:
        logger.info("未加载任何有效自定义headers")
    return custom_headers

# -------------------------- 修改：Selenium获取动态HTML（注入自定义headers） --------------------------
def get_dynamic_html_with_selenium(url, custom_headers):
    logger.info("使用Selenium获取动态渲染的HTML...")
    driver_filename = "chromedriver.exe" if sys.platform.startswith('win') else "chromedriver"
    driver_path = os.path.join(os.getcwd(), driver_filename)
    if not os.path.exists(driver_path):
        logger.error(f"未找到ChromeDriver: {driver_path}")
        if not sys.platform.startswith('win'): logger.error("在macOS/Linux上, 您可能需要运行: chmod +x chromedriver")
        return None
    
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    
    # 1. 处理User-Agent：优先用自定义headers，否则随机默认
    ua = custom_headers.get('User-Agent', random.choice(header_agents))
    chrome_options.add_argument(f"user-agent={ua}")
    
    # 2. 注入其他自定义headers（排除已单独处理的User-Agent）
    for key, value in custom_headers.items():
        if key.lower() == 'user-agent':
            continue
        chrome_options.add_argument(f"--header={key}:{value}")
    logger.debug(f"Selenium注入headers: User-Agent={ua}, 其他headers={json.dumps({k:v for k,v in custom_headers.items() if k.lower()!='user-agent'}, ensure_ascii=False)}")
    
    service = Service(executable_path=driver_path); driver = None
    try:
        driver = webdriver.Chrome(service=service, options=chrome_options); driver.get(url); time.sleep(5)
        html_content = driver.page_source; logger.success("成功使用Selenium获取渲染后的HTML.")
        return html_content
    except Exception as e: logger.error(f"Selenium页面渲染失败: {e}"); return None
    finally:
        if driver: driver.quit()

# -------------------------- 修改：HTTP请求（合并默认头与自定义头，自定义优先） --------------------------
def http_req(url, method='get', custom_headers=None, **kwargs):
    custom_headers = custom_headers or {}  # 避免None值
    kwargs.setdefault('verify', False); kwargs.setdefault('timeout', (20, 60)); kwargs.setdefault('allow_redirects', True)
    
    # 合并headers：默认头 → 自定义头（自定义覆盖默认）
    default_headers = {'User-Agent': random.choice(header_agents)}  # 默认只带User-Agent
    final_headers = {**default_headers, **custom_headers}  # 字典合并，后一个覆盖前一个
    
    # 加入请求参数
    kwargs['headers'] = final_headers
    if SET_PROXY: kwargs['proxies'] = proxies
    
    try:
        logger.debug(f"请求 ({method.upper()}) -> {url} | Headers: {json.dumps(final_headers, ensure_ascii=False)}")
        conn = getattr(requests, method)(url, **kwargs)
        return conn
    except requests.exceptions.RequestException as e: logger.error(f"请求失败 {url}: {e}"); return None

def fill_parameters(parameters):
    filled_params = {"query": {}, "path": {}, "body": {}}
    for param in parameters:
        name = param.get('name'); param_in = param.get('in'); param_type = param.get('type', 'string')
        if not param_in and param.get('paramType'):
            param_in_map = {'path': 'path', 'query': 'query', 'body': 'body'}; param_in = param_in_map.get(param.get('paramType'))
        if not name or not param_in: continue
        value = '1' if param_type in ['integer', 'long', 'number'] else True if param_type == 'boolean' else 'a'
        if param_in in filled_params: filled_params[param_in][name] = value
    return filled_params

# -------------------------- 修改：写入已调用API表格（新增“请求头”列） --------------------------
def output_to_xlsx_called(data):
    try:
        ws_called_apis.append(data); is_success = 200 <= data[6] < 300  # 列索引调整（新增请求头后状态码列变为第7列）
        apply_row_styles(ws_called_apis, ws_called_apis.max_row, fill=SUCCESS_FILL if is_success else None)
    except Exception as e: logger.error(f"写入 '已调用API' sheet页失败: {e}")

def output_to_xlsx_filtered(data):
    try:
        ws_filtered_apis.append(data); apply_row_styles(ws_filtered_apis, ws_filtered_apis.max_row, fill=FILTER_FILL)
    except Exception as e: logger.error(f"写入 '已过滤API' sheet页失败: {e}")

def output_to_xlsx_all(data):
    try:
        ws_all_apis.append(data); fill_color = None
        if data[4] == "Called": fill_color = SUCCESS_FILL
        if data[4] == "Filtered": fill_color = FILTER_FILL
        apply_row_styles(ws_all_apis, ws_all_apis.max_row, fill=fill_color)
    except Exception as e: logger.error(f"写入 '所有API' sheet页失败: {e}")

# -------------------------- 修改：发送请求（传入自定义headers并写入表格） --------------------------
def send_and_process_request(method, request_url, params, summary, source_url, custom_headers):
    # 传入自定义headers到http_req
    response = http_req(
        request_url, 
        method, 
        custom_headers=custom_headers,
        json=params.get('body') if method.lower() == 'post' and params.get('body') else None,
        params=params.get('query') if method.lower() == 'post' and not params.get('body') else params.get('query')
    )
    if response is not None:
        params_for_output = json.dumps(params, ensure_ascii=False, separators=(',', ':'))
        headers_for_output = json.dumps(custom_headers, ensure_ascii=False, separators=(',', ':'))  # 格式化headers为字符串
        resp_text_for_output = ""
        try:
            resp_json = response.json(); resp_text_for_output = json.dumps(resp_json, ensure_ascii=False, separators=(',', ':'))
        except json.JSONDecodeError:
            content_type = response.headers.get('Content-Type', 'unknown').lower(); text_content = response.text
            safe_text_types = ['text/html', 'text/plain', 'application/xml', 'text/xml', 'application/javascript']
            is_text_safe = any(t in content_type for t in safe_text_types)
            if text_content.startswith('PK'):
                resp_text_for_output = f"[二进制内容: 检测到ZIP文件头(PK), Content-Type: {content_type}]"
            elif not is_text_safe:
                resp_text_for_output = f"[非文本内容, Content-Type: {content_type}]"
            else:
                resp_text_for_output = text_content.replace('\n', ' ').replace('\r', '')
        if 200 <= response.status_code < 300: 
            logger.success(f"成功 ({response.status_code}) on [{method.upper()}] {request_url}")
        else: 
            logger.warning(f"失败 ({response.status_code}) on [{method.upper()}] {request_url}")
        
        # 将所有被调用的API都写入表格，无论状态码如何
        output_to_xlsx_called([source_url, method.upper(), request_url, summary, params_for_output, headers_for_output, response.status_code, resp_text_for_output[:32767]])

# -------------------------- 路径清洗工具函数（保留原功能） --------------------------
def clean_path_components(custom_prefix, parsed_path):
    clean_prefix = custom_prefix.rstrip('/')
    if clean_prefix:
        clean_prefix += '/'
    clean_parsed = parsed_path.lstrip('/')
    return f"{clean_prefix}{clean_parsed}" if clean_parsed else clean_prefix

# -------------------------- 修改：V2/V3 解析逻辑（传入自定义headers） --------------------------
def parse_and_scan_v2_v3(data, source_url, args):
    is_v3 = 'openapi' in data; api_version = data.get('openapi' if is_v3 else 'swagger', 'Unknown'); logger.info(f"解析API文档 (版本: {api_version}) 从 {source_url}")
    
    doc_url_parsed = urlparse(source_url)
    domain = f"{doc_url_parsed.scheme}://{doc_url_parsed.netloc}"
    
    base_path = ''
    if is_v3 and data.get('servers'):
        server_obj = data['servers'][0]; server_url = server_obj.get('url', '')
        if '{' in server_url and 'variables' in server_obj:
            for var_name, var_details in server_obj['variables'].items():
                default_value = var_details.get('default', ''); server_url = server_url.replace(f'{{{var_name}}}', str(default_value))
        
        parsed_server_url = urlparse(server_url)
        if not args.force_domain and parsed_server_url.scheme and parsed_server_url.netloc:
            domain = f"{parsed_server_url.scheme}://{parsed_server_url.netloc}"
            logger.info(f"采纳API文档中的服务器地址: {domain}")
        else:
            logger.info(f"强制使用初始URL的服务器地址: {domain}")
            
        base_path = parsed_server_url.path
        
    elif not is_v3: 
        base_path = data.get('basePath', '/')

    definitions = data.get('components', {}).get('schemas', {}) if is_v3 else data.get('definitions', {}); paths = data.get('paths', {})
    if not paths: logger.warning("未在API文档中发现路径."); return
    logger.info(f"发现 {len(paths)} 个路径待测试.")
    for path, methods in paths.items():
        for method, details in methods.items():
            summary = details.get('summary', '无摘要'); temp_params = fill_parameters(details.get('parameters', [])); final_path = path
            for p_name, p_value in temp_params['path'].items():
                placeholder = PATH_VARIABLE_FORMAT.format(param=p_name); final_path = final_path.replace(placeholder, str(p_value))
            
            # 拼接自定义路径前缀
            custom_full_path = clean_path_components(args.custom_path_prefix, final_path)
            full_path = urljoin(base_path + '/', custom_full_path.lstrip('/'))
            request_url = urljoin(domain, full_path)
            
            if method.lower() not in ALLOWED_METHODS:
                reason = f"请求方法 '{method.upper()}' 不在允许列表中"
                output_to_xlsx_all([source_url, method.upper(), request_url, summary, "Filtered"]); output_to_xlsx_filtered([source_url, method.upper(), request_url, summary, reason]); continue
            is_dangerous, reason = is_url_dangerous(request_url, DANGEROUS_KEYWORDS)
            if is_dangerous:
                output_to_xlsx_all([source_url, method.upper(), request_url, summary, "Filtered"]); output_to_xlsx_filtered([source_url, method.upper(), request_url, summary, reason]); continue
            output_to_xlsx_all([source_url, method.upper(), request_url, summary, "Called"])
            param_definitions = []
            for param in details.get('parameters', []): param_definitions.append({'name': param.get('name'), 'in': param.get('in'), 'type': param.get('schema', {}).get('type', 'string')})
            if is_v3 and 'requestBody' in details:
                content = details['requestBody'].get('content', {});
                if 'application/json' in content:
                    schema = content['application/json'].get('schema', {}); ref_name = schema.get('$ref', '').split('/')[-1]
                    if ref_name and ref_name in definitions:
                        for p, d in definitions[ref_name].get('properties', {}).items(): param_definitions.append({'name': p, 'in': 'body', 'type': d.get('type')})
                    elif 'properties' in schema:
                        for p, d in schema.get('properties', {}).items(): param_definitions.append({'name': p, 'in': 'body', 'type': d.get('type')})
            elif not is_v3:
                for param in details.get('parameters', []):
                    if param.get('in') == 'body':
                        schema = param.get('schema', {}); ref_name = schema.get('$ref', '').split('/')[-1]
                        if ref_name and ref_name in definitions:
                            for p, d in definitions[ref_name].get('properties', {}).items(): param_definitions.append({'name': p, 'in': 'body', 'type': d.get('type')})
            params = fill_parameters(param_definitions)
            logger.info(f"测试中: [{method.upper()}] {summary} -> {request_url}")
            # 传入自定义headers到请求函数
            send_and_process_request(method, request_url, params, summary, source_url, args.custom_headers)

# -------------------------- 修改：V1 解析逻辑（传入自定义headers） --------------------------
def parse_and_scan_v1(data, source_url, args):
    logger.info(f"处理 Swagger 1.x 资源列表: {source_url}")
    doc_url_parsed = urlparse(source_url); domain = f"{doc_url_parsed.scheme}://{doc_url_parsed.netloc}"
    base_url_for_discovery = data.get('basePath', source_url)
    for api in data.get('apis', []):
        api_path = api.get('path');
        if not api_path: continue
        declaration_url = urljoin(base_url_for_discovery, api_path.lstrip('/')); logger.info(f"获取V1 API声明从: {declaration_url}");
        # 声明URL请求也携带自定义headers
        decl_res = http_req(declaration_url, custom_headers=args.custom_headers)
        if not decl_res: continue
        try:
            decl_data = decl_res.json(); api_base_path = urlparse(decl_data.get('basePath', '/')).path
            for api_def in decl_data.get('apis', []):
                endpoint_path = api_def.get('path')
                for op in api_def.get('operations', []):
                    method = op.get('method'); summary = op.get('summary', '无摘要'); param_definitions = op.get('parameters', [])
                    params_temp = fill_parameters(param_definitions); final_path = endpoint_path
                    for p_name, p_value in params_temp['path'].items():
                        placeholder = PATH_VARIABLE_FORMAT.format(param=p_name); final_path = final_path.replace(placeholder, str(p_value))
                    
                    # 拼接自定义路径前缀
                    custom_full_path = clean_path_components(args.custom_path_prefix, final_path)
                    full_api_path = (api_base_path.strip('/') + '/' + custom_full_path.strip('/')).replace('//', '/')
                    request_url = urljoin(domain, full_api_path)

                    if not method or method.lower() not in ALLOWED_METHODS:
                        reason = f"请求方法 '{str(method).upper()}' 不在允许列表中"
                        output_to_xlsx_all([source_url, str(method).upper(), request_url, summary, "Filtered"]); output_to_xlsx_filtered([source_url, str(method).upper(), request_url, summary, reason]); continue
                    is_dangerous, reason = is_url_dangerous(request_url, DANGEROUS_KEYWORDS)
                    if is_dangerous:
                        output_to_xlsx_all([source_url, method.upper(), request_url, summary, "Filtered"]); output_to_xlsx_filtered([source_url, method.upper(), request_url, summary, reason]); continue
                    output_to_xlsx_all([source_url, method.upper(), request_url, summary, "Called"])
                    params = fill_parameters(param_definitions)
                    logger.info(f"测试中: [{method.upper()}] {summary} -> {request_url}")
                    # 传入自定义headers到请求函数
                    send_and_process_request(method, request_url, params, summary, source_url, args.custom_headers)
        except (json.JSONDecodeError, AttributeError): logger.error(f"解析或处理V1 API声明失败: {declaration_url}")

def check_url_type(url, custom_headers):
    logger.info(f"检查URL类型: {url}"); 
    # 检查URL时也携带自定义headers
    res = http_req(url, custom_headers=custom_headers)
    if not res: return None, None
    text = res.text; data = None
    try: data = res.json()
    except (json.JSONDecodeError, AttributeError): pass
    if isinstance(data, dict):
        if 'openapi' in data or 'swagger' in data:
            if 'paths' in data: return "api_docs_v2_v3", data
            elif 'apis' in data: return "resource_v1", data
    elif isinstance(data, list) and data and ('location' in data[0] or 'url' in data[0]): return "resource_v2", data
    if '<html' in text.lower():
        url_path = urlparse(url).path
        api_like_keywords = ['api-docs', 'swagger.json', 'openapi.json', 'swagger-resources']
        if any(keyword in url_path for keyword in api_like_keywords):
            logger.warning(f"URL路径 {url_path} 看起来像API端点但返回了HTML，可能是一个错误页面。将中止处理。")
            return None, None
        else: return "html", text
    logger.warning(f"无法确定URL类型: {url}"); return None, None

# -------------------------- 修改：HTML解析（传入自定义headers） --------------------------
def go_swagger_html(url, args):
    logger.info(f"检测到HTML页面. 激活Selenium...")
    # 传入自定义headers到Selenium
    html_content = get_dynamic_html_with_selenium(url, args.custom_headers)
    if not html_content: logger.error(f"获取动态HTML失败 {url}. 中止HTML解析."); return
    api_doc_urls = find_api_docs_aggressively(url, html_content)
    if api_doc_urls:
        logger.success(f"从HTML页面共发现 {len(api_doc_urls)} 个API文档入口.")
        for doc_url in api_doc_urls:
            logger.info(f"--- 开始处理入口: {doc_url} ---"); run(doc_url, args)
    else: logger.error("在HTML页面中未发现任何有效的API定义链接或端点.")

def go_resources(data, source_url, args):
    logger.info(f"处理swagger-resources从 {source_url}")
    if not isinstance(data, list): logger.error("无效的 swagger-resources 格式."); return
    for resource in data:
        if location := resource.get('location') or resource.get('url'):
            docs_url = urljoin(source_url, location)
            logger.info(f"从swagger-resources中发现API文档位置: {docs_url}"); run(docs_url, args)

# -------------------------- 修改：主运行逻辑（传入自定义headers） --------------------------
def run(target_url, args):
    # 检查URL类型时携带自定义headers
    url_type, data = check_url_type(target_url, args.custom_headers)
    if url_type == "api_docs_v2_v3": parse_and_scan_v2_v3(data, target_url, args)
    elif url_type == "resource_v1": parse_and_scan_v1(data, target_url, args)
    elif url_type == "resource_v2": go_resources(data, target_url, args)
    elif url_type == "html": go_swagger_html(target_url, args)
    else: logger.error(f"URL {target_url} 不支持或处理失败.")

def find_api_docs_aggressively(base_url, html_content):
    found_urls = set(); parsed_base_url = urlparse(base_url)
    logger.info("阶段 1: 检查URL查询参数...")
    query_params = parse_qs(parsed_base_url.query)
    config_keys = ['configUrl', 'url', 'urls.primaryName']
    for key in config_keys:
        if key in query_params:
            config_path = query_params[key][0]; config_url = urljoin(base_url, config_path)
            logger.success(f"在URL参数 '{key}' 中直接发现配置URL: {config_url}")
            res = http_req(config_url)
            if res and res.status_code == 200:
                try:
                    config_data = res.json()
                    if 'url' in config_data:
                        api_doc_path = config_data.get('url'); api_doc_url = urljoin(config_url, api_doc_path)
                        logger.success(f"从swagger-config中解析出最终API文档地址: {api_doc_url}"); found_urls.add(api_doc_url); return list(found_urls)
                    elif 'paths' in config_data or 'apis' in config_data:
                         found_urls.add(config_url); return list(found_urls)
                except json.JSONDecodeError:
                    if 'paths' in res.text or 'apis' in res.text:
                        logger.success(f"configUrl {config_url} 直接指向了API文档"); found_urls.add(config_url); return list(found_urls)
    logger.info("阶段 2: 尝试从HTML内容中进行正则匹配...")
    patterns = [r'url\s*:\s*["\']([^"\'<>]+)["\']', r'"swaggerDocUrl"\s*:\s*"([^"\'<>]+)"', r'discoveryUrl\s*:\s*["\']([^"\'<>]+)"', r'["\']([^"\'<>]*?(?:api-docs|swagger\.json|openapi\.json)[^"\'<>]*?)["\']']
    for pattern in patterns:
        matches = re.findall(pattern, html_content, re.IGNORECASE)
        for match in matches:
            if isinstance(match, tuple): match = match[0]
            if match and len(match) > 1 and 'petstore' not in match:
                full_url = urljoin(base_url, match.strip()); found_urls.add(full_url)
    if found_urls:
        logger.success(f"正则匹配成功, 发现 {len(found_urls)} 个潜在URL."); return list(found_urls)
    logger.warning("前两阶段失败, 启动阶段 3: 逐级回溯探测...")
    standard_paths_to_probe = ["/v3/api-docs", "/swagger-resources", "/v2/api-docs", "/api-docs", "/swagger.json", "/openapi.json"]
    path_parts = parsed_base_url.path.strip('/').split('/')
    if path_parts and '.' in path_parts[-1]: path_parts = path_parts[:-1]
    probe_bases = set()
    base_url_root = f"{parsed_base_url.scheme}://{parsed_base_url.netloc}/"
    for i in range(len(path_parts), -1, -1):
        current_path = '/'.join(path_parts[:i])
        probe_bases.add(urljoin(base_url_root, current_path + '/'))
    for base in sorted(list(probe_bases), key=len, reverse=True):
        logger.info(f"--- 以 '{base}' 为基准进行探测 ---")
        for path in standard_paths_to_probe:
            probe_url = urljoin(base, path.lstrip('/'))
            if probe_url in found_urls: continue
            logger.info(f"探测中 -> {probe_url}")
            res = http_req(probe_url)
            if res and res.status_code == 200:
                try:
                    res.json(); logger.success(f"探测成功! 发现API定义入口: {probe_url}"); found_urls.add(probe_url)
                    return list(found_urls)
                except json.JSONDecodeError: logger.warning(f"路径 {probe_url} 可访问但不是有效的JSON, 继续探测...")
    return list(found_urls)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Swagger/OpenAPI通用扫描工具（支持自定义路径前缀+自定义headers）")
    parser.add_argument('-u', '--url', dest='target_url', help='swagger-ui, api-docs, 或 swagger-resources 的URL')
    parser.add_argument('-f', '--file', dest='url_file', help='包含多个目标URL的文件')
    parser.add_argument('--debug', action='store_true', help='在控制台开启Debug日志')
    parser.add_argument('--force-domain', action='store_true', help='强制使用初始URL的域名，忽略API文档内的servers字段')
    parser.add_argument('-cp', '--custom-path-prefix', dest='custom_path_prefix', 
                        default='', help='自定义路径前缀（如 /api），拼接在解析出的接口路径前。示例：-cp /api')
    # -------------------------- 新增：自定义headers参数 --------------------------
    parser.add_argument('-H', '--header', dest='header_list', 
                        action='append', default=[], help='自定义HTTP请求头（支持多个，格式：key:value）。示例：-H token:123456 -H "User-Agent:MyAgent"')
    args = parser.parse_args()
    
    # -------------------------- 解析自定义headers并挂载到args --------------------------
    args.custom_headers = parse_custom_headers(args.header_list)
    
# -------------------------- 保存Excel文件的函数 --------------------------
def save_workbook(base_name="ScanReport"):
    try:
        logger.info("正在进行最后的格式美化，请稍候...")
        auto_adjust_column_width(ws_all_apis); auto_adjust_column_width(ws_called_apis); auto_adjust_column_width(ws_filtered_apis)
        now_time = datetime.now(); filename_time = now_time.strftime("%Y%m%d%H%M")
        output_filename = f'{filename_time}_{base_name}.xlsx'
        workbook.save(output_filename)
        logger.success(f"所有任务完成. 报告已保存至 {output_filename}")
        return output_filename
    except Exception as e:
        logger.error(f"保存Excel文件失败: {e}")
        return None


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Swagger/OpenAPI通用扫描工具（支持自定义路径前缀+自定义headers）")
    parser.add_argument('-u', '--url', dest='target_url', help='swagger-ui, api-docs, 或 swagger-resources 的URL')
    parser.add_argument('-f', '--file', dest='url_file', help='包含多个目标URL的文件')
    parser.add_argument('--debug', action='store_true', help='在控制台开启Debug日志')
    parser.add_argument('--force-domain', action='store_true', help='强制使用初始URL的域名，忽略API文档内的servers字段')
    parser.add_argument('-cp', '--custom-path-prefix', dest='custom_path_prefix', 
                        default='', help='自定义路径前缀（如 /api），拼接在解析出的接口路径前。示例：-cp /api')
    # -------------------------- 新增：自定义headers参数 --------------------------
    parser.add_argument('-H', '--header', dest='header_list', 
                        action='append', default=[], help='自定义HTTP请求头（支持多个，格式：key:value）。示例：-H token:123456 -H "User-Agent:MyAgent"')
    args = parser.parse_args()
    
    # -------------------------- 解析自定义headers并挂载到args --------------------------
    args.custom_headers = parse_custom_headers(args.header_list)
    
    if args.debug: logger.add(sys.stderr, level="DEBUG")
    setup_xlsx_headers()
    try:
        if args.target_url: run(args.target_url, args)
        elif args.url_file:
            try:
                with open(args.url_file, 'r', encoding='utf-8') as f: urls = [line.strip() for line in f if line.strip()]
                for u in urls: run(u, args)
            except FileNotFoundError: logger.error(f"文件未找到: {args.url_file}")
        else: parser.print_help()
    finally:
        base_name = "ScanReport"
        if args.target_url:
            try:
                domain = urlparse(args.target_url).netloc; safe_domain = re.sub(r'[.:\\/*?"<>|]', '_', domain); base_name = safe_domain
            except Exception: base_name = "invalid_url"
        elif args.url_file:
            try:
                 file_basename = os.path.basename(args.url_file); safe_filename = os.path.splitext(file_basename)[0]; base_name = f"from_{safe_filename}"
            except: base_name = "multi_targets"
        save_workbook(base_name)